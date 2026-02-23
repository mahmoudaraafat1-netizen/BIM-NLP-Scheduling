#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
import os

# -----------------------------
# Config
# -----------------------------
STAGE_RULES = {
    "Shuttering":   {"Area": True,  "Volume": False},
    "Steelfixing":  {"Area": False, "Volume": True},
    "Pouring":      {"Area": False, "Volume": True},
    "Deshuttering": {"Area": True,  "Volume": False},
}
STAGES = ["Shuttering", "Steelfixing", "Pouring", "Deshuttering"]

EXPECTED_INPUT_HEADERS = {
    "type":       ["Type", "Activity Type", "Work Type", "نوع النشاط"],
    "element":    ["Element Name", "Element", "Element Type", "العنصر"],
    "area":       ["Area", "مساحة"],
    "volume":     ["Volume", "Qty", "Quantity", "حجم"],
    "total_cost": ["Selling Price Cost", "Total Cost", "Cost", "القيمة", "التكلفة"],
}

OUTPUT_HEADERS = [
    "#", "Activity Name", "Type", "Element", "Stage",
    "Area", "Volume",
    "Total Cost", "Cost %", "Stage Cost"
]


# -----------------------------
# Helpers
# -----------------------------
def normalize(v):
    return "" if v is None else str(v).strip()

def to_float(v, default=None):
    if v is None:
        return default
    s = str(v).strip()
    if not s:
        return default
    try:
        return float(s)
    except Exception:
        return default

def find_header_indices(sheet):
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    name_to_idx = {normalize(h).lower(): i for i, h in enumerate(header_row)}

    def match_one(cands):
        for c in cands:
            k = c.strip().lower()
            if k in name_to_idx:
                return name_to_idx[k]
        return None

    idx = {k: match_one(v) for k, v in EXPECTED_INPUT_HEADERS.items()}
    # fallback للـ 4 الأساسيين لو ناقصين
    for k, fallback_i in zip(["type","element","area","volume"], [0,1,2,3]):
        if idx.get(k) is None:
            idx[k] = fallback_i
    return idx

def autosize(ws):
    for col in ws.columns:
        ml = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            ml = max(ml, len(val))
        ws.column_dimensions[letter].width = min(ml + 2, 60)


# -----------------------------
# Custom Cost Split Dialog
# -----------------------------
class CostSplitDialog(tk.Toplevel):
    """
    نافذة موحّدة لإدخال نسب التوزيع الأربع + عرض المتبقي لحظيًا.
    Buttons:
      - Use (OK)
      - Equal Split
      - Normalize (scale to 100%)
      - Cancel
    """
    def __init__(self, parent, stages):
        super().__init__(parent)
        self.title("Cost Distribution - Concrete Stages")
        self.resizable(False, False)
        self.grab_set()  # modal

        self.stages = stages
        self.vars = {s: tk.StringVar(value="") for s in stages}

        # Grid headers
        tk.Label(self, text="Stage", font=("Arial", 10, "bold")).grid(row=0, column=0, padx=8, pady=6, sticky="w")
        tk.Label(self, text="Percent (%)", font=("Arial", 10, "bold")).grid(row=0, column=1, padx=8, pady=6, sticky="w")

        # Rows for entries
        for i, s in enumerate(self.stages, start=1):
            tk.Label(self, text=s).grid(row=i, column=0, padx=8, pady=4, sticky="w")
            e = tk.Entry(self, textvariable=self.vars[s], width=10, justify="center")
            e.grid(row=i, column=1, padx=8, pady=4)
            # Trace changes to update remainder
            self.vars[s].trace_add("write", lambda *args: self.update_remainder())

        # Remainder label
        self.remainder_var = tk.StringVar(value="Remaining: 100.00%")
        self.remainder_label = tk.Label(self, textvariable=self.remainder_var, fg="blue", font=("Arial", 10, "bold"))
        self.remainder_label.grid(row=len(self.stages)+1, column=0, columnspan=2, padx=8, pady=(8,4), sticky="w")

        # Buttons row
        btn_frame = tk.Frame(self)
        btn_frame.grid(row=len(self.stages)+2, column=0, columnspan=2, pady=10)

        tk.Button(btn_frame, text="Equal Split", command=self.equal_split, width=12).grid(row=0, column=0, padx=6)
        tk.Button(btn_frame, text="Normalize", command=self.normalize_to_100, width=12).grid(row=0, column=1, padx=6)
        tk.Button(btn_frame, text="Use", command=self.on_ok, width=10).grid(row=0, column=2, padx=6)
        tk.Button(btn_frame, text="Cancel", command=self.on_cancel, width=10).grid(row=0, column=3, padx=6)

        self.result = None
        self.update_remainder()

        # center relative to parent
        self.update_idletasks()
        if parent:
            x = parent.winfo_rootx() + (parent.winfo_width() - self.winfo_width()) // 2
            y = parent.winfo_rooty() + (parent.winfo_height() - self.winfo_height()) // 2
            self.geometry(f"+{x}+{y}")

    def get_values(self):
        vals = {}
        for s in self.stages:
            v = self.vars[s].get().strip()
            if v == "":
                continue
            f = to_float(v, default=None)
            if f is not None and f >= 0:
                vals[s] = f
        return vals

    def sum_values(self):
        return sum(self.get_values().values())

    def update_remainder(self):
        s = self.sum_values()
        remaining = 100.0 - s
        # color: red if over, blue otherwise
        if remaining < -1e-9:
            self.remainder_label.config(fg="red")
        else:
            self.remainder_label.config(fg="blue")
        self.remainder_var.set(f"Remaining: {remaining:.2f}%")

    def equal_split(self):
        eq = 100.0 / len(self.stages)
        for s in self.stages:
            self.vars[s].set(f"{eq:.2f}")

    def normalize_to_100(self):
        vals = self.get_values()
        if len(vals) == 0:
            # if all blank -> equal split
            self.equal_split()
            return
        total = sum(vals.values())
        if total <= 0:
            self.equal_split()
            return
        scale = 100.0 / total
        for s in self.stages:
            v = vals.get(s, 0.0) * scale
            self.vars[s].set(f"{v:.2f}")
        self.update_remainder()

    def on_ok(self):
        vals = self.get_values()
        if len(vals) == 0:
            # all blanks -> equal
            self.equal_split()
            vals = self.get_values()
        total = sum(vals.values())

        if abs(total - 100.0) > 1e-6:
            if messagebox.askyesno("Normalize?",
                                   "Entered values don't sum to 100%.\nNormalize automatically to 100%?"):
                # normalize
                self.normalize_to_100()
                vals = self.get_values()
            else:
                # allow user to adjust
                return

        # finalize
        self.result = {s: to_float(self.vars[s].get(), 0.0) for s in self.stages}
        self.destroy()

    def on_cancel(self):
        self.result = None
        self.destroy()


def ask_cost_distribution(parent):
    """
    يفتح Dialog يسألك لو عايز توزع، ولو وافقت يفتح نافذة موحّدة فيها الأربع مراحل والمتبقي.
    Returns: (distribute: bool, pct_dict: dict or None)
    """
    distribute = messagebox.askyesno("Cost Distribution",
                                     "Do you want to distribute cost across concrete stages?")
    if not distribute:
        return False, None

    dlg = CostSplitDialog(parent, STAGES)
    parent.wait_window(dlg)
    if dlg.result is None:
        # user canceled on the split dialog
        return False, None

    return True, dlg.result


# -----------------------------
# Core logic
# -----------------------------
def build_activity_list(input_path, distribute_cost=False, pct_dict=None, save_path=None):
    wb_in = openpyxl.load_workbook(input_path)
    sh = wb_in.active
    idx = find_header_indices(sh)

    wb_out = Workbook()
    ws = wb_out.active
    ws.title = "Activity_List"

    for i, h in enumerate(OUTPUT_HEADERS, 1):
        ws.cell(row=1, column=i, value=h)

    out_row = 2
    counter = 1

    for row in sh.iter_rows(min_row=2, values_only=True):
        t = normalize(row[idx["type"]]) if len(row) > idx["type"] else ""
        e = normalize(row[idx["element"]]) if len(row) > idx["element"] else ""
        a = row[idx["area"]]   if len(row) > idx["area"]   else None
        v = row[idx["volume"]] if len(row) > idx["volume"] else None
        c = row[idx["total_cost"]] if idx.get("total_cost") is not None and len(row) > idx["total_cost"] else None
        total_cost = to_float(c, default=0.0)

        if not (t or e or a or v or total_cost):
            continue

        is_concrete = "concrete" in t.lower()

        if is_concrete:
            for stage in STAGES:
                rules = STAGE_RULES.get(stage, {"Area": False, "Volume": False})

                ws.cell(out_row, 1, counter)
                ws.cell(out_row, 2, f"{t} - {stage} - {e}")
                ws.cell(out_row, 3, t)
                ws.cell(out_row, 4, e)
                ws.cell(out_row, 5, stage)

                ws.cell(out_row, 6, a if rules["Area"] else None)
                ws.cell(out_row, 7, v if rules["Volume"] else None)

                if distribute_cost and pct_dict:
                    pct = (pct_dict.get(stage, 0.0) or 0.0) / 100.0  # fraction
                    stage_cost = round(pct * total_cost, 2)

                    ws.cell(out_row, 8, total_cost)      # Total Cost
                    ws.cell(out_row, 9, pct)             # Cost % (fraction)
                    ws.cell(out_row, 9).number_format = '0.00%'
                    ws.cell(out_row, 10, stage_cost)     # Stage Cost (rounded)
                    # ws.cell(out_row, 10).number_format = '#,##0.00'  # uncomment if you want fixed 2 decimals
                else:
                    if stage == "Pouring":
                        ws.cell(out_row, 8, total_cost)
                    ws.cell(out_row, 9, None)
                    ws.cell(out_row, 10, None)

                out_row += 1
                counter += 1

        else:
            ws.cell(out_row, 1, counter)
            ws.cell(out_row, 2, f"{e} - {t}")
            ws.cell(out_row, 3, t)
            ws.cell(out_row, 4, e)
            ws.cell(out_row, 5, None)
            ws.cell(out_row, 6, None)
            ws.cell(out_row, 7, None)
            ws.cell(out_row, 8, total_cost)
            ws.cell(out_row, 9, 1.0)
            ws.cell(out_row, 9).number_format = '0.00%'
            ws.cell(out_row, 10, round(total_cost, 2))
            # ws.cell(out_row, 10).number_format = '#,##0.00'

            out_row += 1
            counter += 1

    ws.freeze_panes = "A2"
    autosize(ws)

    if save_path is None:
        default_name = f"Activity_List_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        save_path = filedialog.asksaveasfilename(
            title="Save Activity List As",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=default_name
        )
        if not save_path:
            messagebox.showinfo("Cancelled", "Save operation cancelled.")
            return None

    wb_out.save(save_path)
    return save_path


# -----------------------------
# Run
# -----------------------------
if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()

    # 1) Select input file
    input_path = filedialog.askopenfilename(
        title="Select Input Excel File",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    if not input_path:
        messagebox.showinfo("Cancelled", "No input file selected.")
        raise SystemExit

    # 2) Ask cost distribution with unified dialog
    distribute, pct = ask_cost_distribution(root)

    # 3) Build & Save
    out_path = build_activity_list(
        input_path,
        distribute_cost=distribute,
        pct_dict=pct,
        save_path=None
    )

    if out_path:
        messagebox.showinfo("Success", f"Activity List saved:\n{out_path}")
        print(f"[OK] Saved: {out_path}")
