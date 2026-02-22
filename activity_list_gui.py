import openpyxl
from openpyxl import Workbook
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox

def generate_activity_list_from_excel(input_file):
    try:
        workbook = openpyxl.load_workbook(input_file)
        sheet = workbook.active
    except FileNotFoundError:
        messagebox.showerror("Error", f"The file '{input_file}' was not found.")
        return
    except Exception as e:
        messagebox.showerror("Error", f"Error reading the Excel file: {e}")
        return

    # إنشاء ملف جديد لكتابة النتائج
    output_workbook = Workbook()
    ws_output = output_workbook.active
    ws_output.title = f"Activity_List_{datetime.now().strftime('%H%M%S')}"

    # تعريف مراحل النشاط
    activity_stages = ["Steelfixing", "Shuttering", "Pouring", "Deshuttering"]

    # كتابة العناوين في الملف الجديد
    headers = ["#", "Activity Name", "Area", "Volume"]
    for col_num, header in enumerate(headers, 1):
        ws_output.cell(row=1, column=col_num, value=header)

    # قراءة البيانات من الملف المدخل
    output_row = 2
    for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True):  # Skip header row
        activity_type, element_type, area_value, volume_value = row

        if "Concrete" in activity_type:
            # تقسيم النشاط إلى مراحل
            for stage in activity_stages:
                ws_output.cell(row=output_row, column=1).value = output_row - 1  # Activity number
                ws_output.cell(row=output_row, column=2).value = f"{activity_type} - {stage} - {element_type}"

                # توزيع المساحة والحجم
                if stage == "Pouring":
                    ws_output.cell(row=output_row, column=3).value = None
                    ws_output.cell(row=output_row, column=4).value = volume_value
                elif stage in ["Shuttering", "Deshuttering"]:
                    ws_output.cell(row=output_row, column=3).value = area_value
                    ws_output.cell(row=output_row, column=4).value = None
                else:
                    ws_output.cell(row=output_row, column=3).value = None
                    ws_output.cell(row=output_row, column=4).value = None

                output_row += 1
        else:
            # لو النوع مش Concrete
            ws_output.cell(row=output_row, column=1).value = output_row - 1
            ws_output.cell(row=output_row, column=2).value = f"{element_type} - {activity_type}"
            ws_output.cell(row=output_row, column=3).value = None
            ws_output.cell(row=output_row, column=4).value = None
            output_row += 1

    # حفظ الملف الجديد
    output_file = f"Activity_List_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_workbook.save(output_file)
    messagebox.showinfo("Success", f"The Activity List has been generated as '{output_file}'!")

def select_file():
    input_file = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    if input_file:
        generate_activity_list_from_excel(input_file)

# تصميم الواجهة الرسومية
root = tk.Tk()
root.title("Activity List Generator")

# زر لاختيار ملف الإكسل
btn_select_file = tk.Button(root, text="Select Excel File", command=select_file, width=20, height=2)
btn_select_file.pack(pady=20)

# تشغيل التطبيق
root.mainloop()


